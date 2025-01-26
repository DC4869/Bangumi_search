import json
import openpyxl as xl
import random
import requests
import time
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill

def get_bangumi_link(keyword, type, maxNumber, year):
    keyword = keyword.replace('（第1期）', '')
    keyword = keyword.replace('（第2期）', '')
    keyword = keyword.replace('（第3期）', '')
    keyword = keyword.replace('（第4期）', '')
    keyword = keyword.replace('（第5期）', '')
    keyword = keyword.replace('!', '')
    keyword = keyword.replace('/', '')
    # SubjectType: 1 Book, 2 Anime, 3 Music, 4 Game, 6 Sanjigen
    url = f'https://api.bgm.tv/search/subject/{keyword}?type={type}&responseGroup=large&max_results={maxNumber}'
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11'
        #'Accept' : '*/*',
        #'Accept-Language': 'en-US,en,ja,zh,cn;q=0.8',
        #'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82',
    }
    print(url)
    time.sleep(random.uniform(0., 1.))
    response = requests.get(url, headers=headers)
    #print(response.status_code)
    bgm_id = ''
    bgm_link = ''
    bgm_name = ''
    bgm_name_cn = ''
    if response.status_code == 200:
        if 'json' in response.headers.get('Content-Type'):
            html = response.json()
            if html['list'] is not None:
                min_number = min(len(html['list']), maxNumber)
                for i in range(0, min_number):
                    air_date = html['list'][i]['air_date']
                    air_year = air_date[0:4]
                    air_month = air_date[5:7]
                    if air_year == year:
                        bgm_id = html['list'][i]['id']
                        bgm_link = html['list'][i]['url']
                        bgm_name = html['list'][i]['name']
                        bgm_name_cn = html['list'][i]['name_cn']
                        break

    return bgm_id, bgm_link, bgm_name, bgm_name_cn

def get_bangumi_collection_status(user_id, bgm_id):
    url = f'https://api.bgm.tv/v0/users/{user_id}/collections/{bgm_id}'
    headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11'}
    time.sleep(random.uniform(0., 1.))
    response = requests.get(url, headers=headers)
    collection_status = 0
    if response.status_code == 200:
        if 'json' in response.headers.get('Content-Type'):
            html = response.json()
            if html['type'] is not None:
                # CollectionType: 1 Wish, 2 Collect, 3 Doing, 4 On_hold, 5 Dropped
                collection_status = html['type']

    return collection_status

#Years = ['2000', '2001', '2002', '2003', '2004', '2005', '2006', '2007', '2008', '2009', '2010', '2011', '2012', '2013', '2014', '2015', '2016', '2017', '2018', '2019', '2020', '2021', '2022', '2023', '2024']
Years = ['2004']
for year in Years:
    input_xlsx = 'Index/Index_{}.xlsx'.format(year)
    wb = xl.load_workbook(input_xlsx)
    sheets = wb.sheetnames
    for sheet_name in sheets:
        worksheet = wb[sheet_name]
        for row in range(3,worksheet.max_row+1):
            #for column in 'B':  #Here you can add or reduce the columns
            cell_name = '{}{}'.format('B', row)
            anime_name = worksheet[cell_name].value # the value of the specific cell
            print(anime_name) 
            bgm_id, bgm_link, bgm_name, bgm_name_cn = get_bangumi_link(anime_name, 2, 10, year)
            print(bgm_id, bgm_link, bgm_name, bgm_name_cn)
            
            bgm_id_cell = '{}{}'.format('I', row)
            bgm_link_cell = '{}{}'.format('J', row)
            bgm_name_cell = '{}{}'.format('K', row)
            bgm_name_cn_cell = '{}{}'.format('L', row)
        
            worksheet[bgm_id_cell] = bgm_id
            worksheet[bgm_link_cell] = bgm_link
            worksheet[bgm_name_cell] = bgm_name
            worksheet[bgm_name_cn_cell] = bgm_name_cn

            bgm_collect_status = get_bangumi_collection_status(416773, bgm_id)
            bgm_status_cell = '{}{}'.format('O', row)
            if bgm_collect_status == 1:
                worksheet[bgm_status_cell] = 'Wish'
                worksheet[bgm_status_cell].fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
            elif bgm_collect_status == 2:
                worksheet[bgm_status_cell] = 'Finish'
                worksheet[bgm_status_cell].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            elif bgm_collect_status == 3:
                worksheet[bgm_status_cell] = 'Ongoing'
                worksheet[bgm_status_cell].fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
            elif bgm_collect_status == 4:
                worksheet[bgm_status_cell] = 'On hold'
                worksheet[bgm_status_cell].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            elif bgm_collect_status == 5:
                worksheet[bgm_status_cell] = 'Dropped'
                worksheet[bgm_status_cell].fill = PatternFill(start_color='FF00FF', end_color='FF00FF', fill_type='solid')
            else:
                worksheet[bgm_status_cell] = 'Null'
                worksheet[bgm_status_cell].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    
    output_xlsx = 'output/Index_{}_bgm.xlsx'.format(year)
    wb.save(output_xlsx)
